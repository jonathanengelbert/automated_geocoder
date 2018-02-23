# -*- coding: utf-8 -*-

#IMPERFECT ADDRESSES
#
#Description: This script standardizes addresses provided in an excel workbook
#Last Modified: 02/02/2018
###############################################################################


#FUNCTIONALITY:
#
# 1. Tranformations:
#
# --> Address types are abbreviated, following the standards found here:
#     https://data.sfgov.org/Geographic-Locations-and-Boundaries/Street-Names/6d9h-4u5v/data
# --> Everything to the right of an address type is wiped out (as in
#     "123 1st street, #Apt 345"  --> "123 1st st"
# --> Apostrophes are removed and leave no whitespace in between characters
# --> Ranges (as in "517-520 1st Street") are eliminated. Only the first
# number in the range must is kept
# --> No fractions in addresses (as in 10/2 Market Street)
# --> Streets and avenues from 1-9 must have a "0" added to the left (as in
# 7th street --> 07th street
# --> Cardinal directions are kept to the right of the address type in 13
# cases.
#
#
# 2. Processes:
#
# --> Book to be transformed is loaded as object "wb"
# --> Output book is created as object "wb2"
# --> Active sheet is stored to object "ws"
# --> First column is copied and pasted into the second column of "ws"
# --> Abbreviates addressees and wipes out characters to the right of
#     address type
# --> Makes sure that cardinal directions are kept (as in 123 1st NE)
# --> Remove apostrophes, periods, octothorpe and commas
# --> Removes ranges
# --> Removes letters mixed with numbers
# --> Removes fractions in addresses
# --> Adds a number to single digit addresses (as in 7th ave --> 07th Ave)
# --> Transforms BAYSHORE BLVD --> BAY SHORE BLVD
# --> Handles Embarcadero Center Transformations as:
#-------------1 embarcadero center --> 301 CLAY ST
#-------------2 embarcadero center –-> 201 CLAY ST
#-------------3 embarcadero center –-> 101 CLAY ST
#-------------4 embarcadero center –-> 150 DRUMM ST
# --> Handles addresses commonly entered without address type
# --> Writes new header for column with value addresses
# --> Saves workbook changes



import openpyxl
from openpyxl.utils import get_column_letter
import re
from numerate_xl import get_address_column


#Loads the input workbook. Note that the workbook has its title harcoded:
wb = openpyxl.load_workbook('test.xlsx')
# Assigns the active sheet of the workbook
ws = wb.active
#Assigns the values retrived from cells to variable "value:
value = (ws.cell)

#Assigns variable to last column of workbook sheet(where transformations
# will be written)

last_column = ws.max_column +1
last_column_letter = get_column_letter(last_column)
#Finds column with addresses to be transformed:
target = get_address_column(wb)

#------------------------------------------------------------------------------

def transform():

    print("\nProcessing Addresses...")
    for i in range(1, ws.max_row + 1):
             #Retrieves address column target from input workbook
             value = (ws.cell(row=i, column=target).value)
             #Copies original address column into memory for processing
             ws.cell(row=i, column=target).value = value
             if(value):
                    value = (ws.cell(row=i, column=target).value).lower()

#------------------------------------------------------------------------------

                #ADDRESS TYPE TRANSFORMATIONS

                    # 1. If address type is long, it transforms it into short type
                    # 2. If address type is short, it drops all characters to the right of address type

                    # Alley:

                    value = re.sub("(?<![0-9\s]) alley.*$", ' aly', value)
                    value = re.sub(" aly[\s|.|,].*$", ' aly', value)

                    #Avenue:

                    # HANDLES CASES CONTAINING CADINAL DIRECTIONS TO THE RIGHT OF ADDRESS TYPE:

                            #Looks for addresses in 25th ave north

                    north25 = re.search("25(th)?\sav(e)?(nue)?\sn(orth)?", value)
                    if north25:
                            value = re.sub("[^0-9].*$", " 25th ave north", value)

                            #Looks for addresses in Buena Vista Ave east

                    bv_east = re.search("buena\svista\sav(e)?(nue)?\se(ast)?", value)
                    if bv_east:
                            value = re.sub("[^0-9].*$", " buena vista ave east", value)
                            # Looks for addresses in Buena Vista Ave west

                    bv_west = re.search("buena\svista\sav(e)?(nue)?\sw(est)?",
                                                value)
                    if bv_west:
                            value = re.sub("[^0-9].*$", " buena vista ave west",
                                                   value)

                            # Looks for addresses in Burnett Ave north

                    burnett_north = re.search("burnett\sav(e)?(nue)?\sn(orth)?",
                                                value)
                    if burnett_north:
                            value = re.sub("[^0-9].*$", " burnett ave north",
                                                   value)

                            # Looks for addresses in South Van Ness Ave

                    s_van_ness = re.search("((s)(outh)?\svan\sness\s(ave)(nue)?)",value)

                    if s_van_ness:
                        value = re.sub("\s(s)(outh)?\s.*$", " south van ness ave", value)

                            # Looks for addresses in West View Ave

                    w_view_ave = re.search("((w)(est)?\sview\s(ave)(nue)?)",
                                           value)

                    if w_view_ave:
                        value = re.sub("\s(w)(est)?\s.*$", " west view ave", value)

                    # Handles Addresses in Avenues with letter names (
                    # (Ex: 23 Avenue B)

                    letter_named_avenues = re.search("\s(ave)?(avenue)?\s["
                                                     "b|c|d|e|f|g|h|i|m|n]",
                                                     value)

                    if letter_named_avenues:
                        value = re.sub("(avenue)?(ave)?\s(["
                                       "b|c|d|e|f|g|h|i|m|n]\s).*$", "avenue \\3",
                                       value)

                    else:
                            if "25th ave north" not in value and "buena vista ave east" \
                                    not in value and "buena vista ave west" not in value\
                                    and "burnett ave north" not in value and \
                                    "ave of the palms" not in value and "avenue " \
                                                                        "b" not \
                                    in value and "avenue c"\
                                    not in value and "avenue d" not in value and\
                                           "avenue e" not in value and "avenue " \
                                                                       "f" not \
                                           in value and "avenue g" not in value \
                                           and "avenue h" not in value and \
                                           "avenue i" not in value and value and\
                                           "avenue m" not in value and "avenue " \
                                                                       "n" not \
                                           in value:
                                    value = re.sub("(?<![0-9\s]) avenue.*$", ' ave', value)
                                    value = re.sub(" av[e]?[\s|.|,].*$", ' ave', value)



                    #Boulevard:

                            # Looks for addresses in South Hill Blvd

                            south_hill_blvd = re.search(
                                "((\ss(outh)?)\shill\s(blvd)?("
                                "boulevard)?)",
                                value)

                            if south_hill_blvd:
                                value = re.sub("[^0-9].*$",
                                               " south hill blvd", value)

                           # Looks for addresses in Lake Merced Blvd

                    lake_merced_blvd = re.search("\slake\smerced\s["
                                                 "boulevard|blvd]",value)
                    if lake_merced_blvd:
                            value = re.sub("[^0-9\s].*$", "lake merced blvd",
                                           value)

                           #Looks for addresses in Mission Bay Boulevard south

                    mission_south = re.search("\smission\sbay\s(boulevard)?("
                                              "blvd)?\s(s)(outh)?",value)
                    if mission_south:
                            value = re.sub("[^0-9].*$", " mission bay blvd south", value)

                           #Looks for addresses in Mission Bay Boulevard north

                    mission_north = re.search("\smission\sbay\s(boulevard)?("
                                              "blvd)?\s(n)("
                                              "orth)?",
                                              value)
                    if mission_north:
                            value = re.sub("[^0-9].*$", " mission bay blvd north", value)

                    else:
                        if "mission bay blvd south" not in value and "mission " \
                                                                     "bay blvd " \
                                                                     "north not " \
                                                                     "in value":
                            value = re.sub("(?<![0-9\s]) boulevard.*$", ' blvd',
                                           value)
                            value = re.sub("blvd[\s|.|,].*$", 'blvd', value)

                    #Circle:

                    value = re.sub("(?<![0-9\s]) circle.*$", ' cir', value)
                    value = re.sub(" cir[\s|.|,].*$", ' cir', value)

                    #Court:

                            # Looks for addresses in North View Court

                    n_view_ct = re.search("((n)(orth)?\sview\s(ct)?(court)?)",
                                         value)

                    if n_view_ct:
                        value = re.sub("\s(n)(orth)?\s.*$", " north view ct",
                                       value)


                    value = re.sub("(?<![0-9\s]) court.*$", ' ct', value)
                    value = re.sub(" ct[\s|.|,].*$", ' ct', value)


                    #Drive:

                    if "carlton" not in value:

                        value = re.sub("(?<![0-9\s]) drive.*$", ' dr', value)
                        value = re.sub(" dr[\s|.|,].*$", ' dr', value)

                    #Expressway:

                    value = re.sub("(?<![0-9\s]) expressway.*$", ' expy', value)
                    value = re.sub(" expy[\s|.|,].*$", ' expy', value)

                    #Highway:

                    value = re.sub("(?<![0-9\s]) highway.*$", ' hwy', value)
                    value = re.sub(" hwy[\s|.|,].*$", ' hwy', value)

                    #Hill:

                    if "lake merced hill" not in value:

                        value = re.sub("(?<![0-9])[a-z]?\s\shill.*$", 'hl',
                                           value)
                        value = re.sub(" hl[\s|.|,].*$", ' hl', value)


                    #Lane:


                    value = re.sub("(?<![0-9\s]) lane\s.*$", ' ln', value)
                    value = re.sub(" ln[\s|.|,].*$", ' ln', value)

                    #Loop:

                    value = re.sub("(?<![0-9\s]) loop.*$", ' loop', value)


                    #Park:

                    if "street" not in value and "st" not in value\
                    and "circle" not in value and "cir" not in value\
                    and "boulevard" not in value and "blvd" not in value\
                    and "road" not in value and "rd" not in value\
                    and "drive" not in value and "dr" not in value\
                    and "avenue" not in value and "ave" not in value\
                    and "lane" not in value and "ln" not in value:


                        value = re.sub("(?<![0-9\s]) park\s.*$", ' park', value)

                            # Looks for addresses in South Park

                    s_park = re.search("\s(s)(outh)?\spark", value)

                    if s_park:
                        value = re.sub("\ss\spark.*$", ' south park', value)


                    #Place:

                            # Adds "DR" to Dr Carlton B Goodlet Pl

                    dr_carlton_pl = re.search("carlton", value)
                    if dr_carlton_pl:
                        value = re.sub("(dr\s)?carlton", "dr carlton", value)

                    value = re.sub("(?<![0-9\s]) place.*$", ' pl', value)
                    value = re.sub("(\s)pl[\s|.|,|-].*$", ' pl', value)

                    #Plaza:

                    value = re.sub("(?<![0-9\s]) plazza.*$", ' plz', value)
                    value = re.sub("(\s)plz[\s|.|,|-].*$", ' plz', value)

                    #Pier:

                    value = re.sub("([0-9\s])pier\s.*$", ' pier', value)

                    #Road:

                            #Looks for addresses in West Point Rd

                    west_point = re.search("((\sw(est)?)\spoint\s(rd)?("
                                            "road)?|point\s(rd)?(road)?("
                            "\sw(est)?))",
                            value)

                    if west_point:
                        value = re.sub("[^0-9].*$",
                                       " west point rd", value)

                    value = re.sub("(?<![0-9\s]) road.*$", ' rd', value)
                    value = re.sub(" rd[\s|.|,].*$", ' rd', value)

                    #Steps:

                    value = re.sub("(?<![0-9\s]) steps.*$", ' stps', value)
                    value = re.sub(" stps[\s|.|,].*$", ' stps', value)

                    #Street:

                            #Looks for addresses in West Clay Street

                    west_clay = re.search("((\sw(est)?)\sclay\sst("
                                        "reet)?|clay\sst(reet)?("
                            "\sw(est)?))",
                            value)

                    if west_clay:
                        value = re.sub("[^0-9].*$",
                                       " west clay st", value)

                            # Looks for addresses in Lake Merced Hill St south

                    lake_merced_south = re.search("(\s(s)("
                                                "outh)?\slake\smerced\s(hl)?("
                                                "hls)?(hill)?"
                                                "|\slake\smerced\s(hl\s)?(hls\s)?("
                                                "hill\s)?(st)(reet)?\s("
                                                "s)(outh)?)",
                                                value)
                    if lake_merced_south:
                            value = re.sub("[^0-9].*$"," lake merced hill st south", value)

                            # Looks for addresses in Lake Merced Hill St north

                    lake_merced_north = re.search("(\s(n)("
                                                "orth)?\slake\smerced\s(hl)?("
                                                "hls)?(hill)?"
                                                "|\slake\smerced\s(hl\s)?(hls\s)?("
                                                "hill\s)?(st)(reet)?\s("
                                                "n)(orth)?)",
                                                value)
                    if lake_merced_north:
                            value = re.sub("[^0-9].*$"," lake merced hill st north", value)


                            #Looks for addresses in North Point Street

                    north_point = re.search("((\sn(orth)?)\spoint\sst(reet)?|point\sst(reet)?("
                            "\sn(orth)?))",
                            value)

                    if north_point:
                        value = re.sub("[^0-9].*$",
                                       " north point st", value)

                            # Looks for addresses in Willard St north

                    willard_north = re.search("((\sn(orth)?)\swillard\sst("
                                              "reet)?|willard\sst(reet)?("
                            "\sn(orth)?))",
                            value)

                    if willard_north:
                            value = re.sub("[^0-9].*$",
                                           " willard st north", value)

                    else:
                            if "lake merced hill st south" not in value and "lake merced hill st north" not in value and "willard st north" not in value:
                                    value = re.sub("(?<![0-9\s]) street(-)?.*$",
                                                   ' st', value)
                                    value = re.sub(" st[\s|.|,|-].*$", ' st',
                                                   value)
                    #Square

                    if "columbia"in value:
                        value = re.sub("(?<![0-9\s]) square.*$", ' square st',
                                       value)
                        value = re.sub(" sq.*$", ' square st', value)

                    else:
                        value = re.sub("(?<![0-9\s]) square.*$", ' sq', value)
                        value = re.sub(" sq.*$", ' sq', value)

                    #Terrace:

                            # Looks for addresses in West Crystal Cove

                    w_crystal_cove = re.search("(\s(w)(est)?\scrystal\scove)",
                                               value)

                    if w_crystal_cove:
                        value = re.sub("(\s(w)(est)?).*$", " west crystal cove ter"
                                       , value)


                            # Looks for addresses in East Crystal Cove

                    e_crystal_cove = re.search("(\s(e)(ast)?\scrystal\scove)",
                                               value)

                    if e_crystal_cove:
                        value = re.sub("(\s(e)(ast)?).*$", " east crystal cove ter"
                                       , value)


                    value = re.sub("(?<![0-9\s]) terrace.*$", ' ter', value)
                    value = re.sub(" ter\s.*$", ' ter', value)

                    #Tunnel:

                    if "ave" not in value and "avenue" not in value:
                        value = re.sub("(?<![0-9\s]) tunnel.*$", ' tunl', value)
                        value = re.sub(" tunnel.*$", ' tunl', value)

                    #Way:

                    value = re.sub("(?<![0-9\s]) way[\s|.|,|-].*$", ' way', value)



                #REMOVES APOSTROPHES, PERIODS, OCTOTHORPE  AND COMMAS

                    value = value.replace("'", "")
                    value = value.replace(".", "")
                    value = value.replace(",", "")
                    value = value.replace("#", "")
                #REMOVES RANGES

                    value = re.sub("(?<=[0-9])(\s)?-(\s)?\d*", '', value)

                #REMOVES LETTERS MIXED WITH NUMBERS

                    value = re.sub("(?<=\d)(\s)?[a-z]\s", " ", value)

                #REMOVES FRACTIONS IN ADDRESSES

                    value = re.sub("[0-9]*/[0-9]\s*",'', value)

                #ADDS A NUMBER TO SINGLE DIGIT ST AND AVE ADDRESS TYPES

                    value = re.sub("\s(?=1st|2nd|3rd|4th|5th|6th|7th|8th|9th\s)",
                                   ' 0', value)

                #HANDLES BAYSHORE BLVD TRANSFORMATION:

                    value = re.sub("bayshore", "bay shore", value)

                #HANDLES EMBARCADERO CENTER TRANSFORMATIONS:

                    value = re.sub("1\sembarcadero\s(ctr)?(center)?.*$",
                                   "301 clay st",
                                   value)
                    value = re.sub("2\sembarcadero\s(ctr)?(center)?.*$",
                                   "201 clay st", value)
                    value = re.sub("3\sembarcadero\s(ctr)?(center)?.*$",
                                   "101 clay st", value)
                    value = re.sub("4\sembarcadero\s(ctr)?(center)?.*$",
                                   "150 drumm st", value)
                    value = re.sub("5\sembarcadero\s(ctr)?(center)?.*$",
                                   "22 drumm st", value)

                #HANDLES ADDRESSES COMMONLY ENTERED WITHOUT STREET TYPE:

                    # Broadway

                    if "broadway" in value:
                        value = re.sub("broadway.*$", "broadway", value)

                    #California Street

                    if "california" in value:
                        value = re.sub("california.*$", "california st", value)

                    #Cesar Chavez

                    if "cesar chavez" in value:
                        value = re.sub("cesar chavez.*$", "cesar chavez st", value)

                    #Lombard Street

                    if "lombard" in value:
                        value = re.sub("lombard.*$", "lombard st", value)

                # HANDLES "THE EMBARCADERO" CASES:

                    if "the embarcadero" in value:
                        value = re.sub("the embarcadero.*$", "the embarcadero",
                                       value)

                # HANDLES FORT MASON ADDRESSES:

                    fort_mason = re.search("fort mason", value)
                    if fort_mason:
                        value = re.sub("fort mason.*$", "fort mason", value)

                # DROPS "AVE" FROM AVENUE OF THE PALMS:

                    ave_of_the_palms = re.search("(ave)?(avenue)? of the palms",
                                                 value)
                    if ave_of_the_palms:
                        value = re.sub("((ave)?(avenue)?\sof\sthe\spalms\s).*$",
                                             "avenue of the palms",
                                             value)

                # DROPS STRING TO THE RIGHT OF "LA AVANZADA:

                    la_avanzada = re.search("la avanzada", value)
                    if la_avanzada:
                        value = re.sub("la avanzada.*$", "la avanzada", value)

                #COPIES ALL TRANSFORMATIONS INTO NEW COLUMN)

                    ws.cell(row=i, column=last_column).value = value


#------------------------------------------------------------------------------
    #CLOSES AND SAVES OUTPUT WORKBOOK

    # Writes header of new column

    ws.cell(row=1, column=last_column).value = "transformed_address"

    #Saves the workbook
    wb.save('transformed.xlsx')
    
#------------------------------------------------------------------------------
    #CONDITIONAL FOR EXTERNAL EXECUTION

if __name__ == "__main__":
    transform()
    print("\nADDRESSES SUCCESSFULLY PROCESSED")
#------------------------------------------------------------------------------




